"""
LA27 Outreach Agent v2 — with AI personalization
- Gmail API primary sender, Brevo fallback
- AI opener via Claude Haiku when on AC power
- Language detection by email domain + location + company
"""

import csv
import time
import json
import os
import re
import datetime
import requests
import openpyxl
from pathlib import Path
from dotenv import load_dotenv

BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / ".env")

BREVO_API_KEY_A = os.getenv("BREVO_API_KEY")
BREVO_API_KEY_B = os.getenv("BREVO_API_KEY_B")
BREVO_API_KEY_C = os.getenv("BREVO_API_KEY_C")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY")
CRM_PATH = Path(os.getenv("CRM_PATH", BASE_DIR.parent / "LA27_CRM.xlsx"))
LEADS_CSV = CRM_PATH.parent / "LA27_leads_with_email.csv"
LOG_PATH = Path(os.getenv("LOG_PATH", BASE_DIR / "logs"))
DAILY_LIMIT = 870  # Total combined limit (290 * 3)
SENDER_EMAIL = os.getenv("SENDER_EMAIL", "tim@la27productions.com")
SENDER_NAME = os.getenv("SENDER_NAME", "Tim | LA 27 Productions")
EMAIL_TEMPLATES = {
    "en": [
        {
            "subject": "Sound for {company}'s next campaign",
            "body": """Hi,

LA 27 Productions is a premium music studio based in Barcelona. We compose 100% original music for brands, agencies and film — built from scratch with live instruments, professional recording and zero stock libraries. Every piece is exclusive to your project. No royalties, no shared licenses, no music your competitors are already using.

Hear our work at la27productions.com

Would you be open to a quick call if this fits any upcoming projects?

Tim Helmes
Founder & Music Director"""
        }
    ],
    "es": [
        {
            "subject": "Sonido para la próxima campaña de {company}",
            "body": """Hola,

LA 27 Productions es un estudio de música premium en Barcelona. Componemos música 100% original para marcas, agencias y film — grabada con instrumentos reales, producción profesional y sin librerías de stock. Cada pieza es exclusiva para tu proyecto. Sin royalties, sin licencias compartidas.

Escucha nuestro trabajo en la27productions.com

¿Estarías abierto a una llamada rápida si esto encaja con algún proyecto?

Tim Helmes"""
        }
    ],
    "de": [
        {
            "subject": "Sound für {company}s nächste Kampagne",
            "body": """Hallo,

LA 27 Productions ist ein Premium-Musikstudio mit Sitz in Barcelona. Wir komponieren 100% Originalmusik für Marken, Agenturen und Film — von Grund auf mit echten Instrumenten, professioneller Aufnahme und ohne Stock-Bibliotheken. Jedes Stück ist exklusiv für Ihr Projekt. Keine Lizenzgebühren, keine geteilten Lizenzen, keine Musik, die Ihre Konkurrenz bereits nutzt.

Hören Sie unsere Arbeiten auf la27productions.com

Wären Sie offen für einen kurzen Call, falls das zu aktuellen Projekten passt?

Tim Helmes"""
        }
    ],
    "fr": [
        {
            "subject": "Le son pour la prochaine campagne de {company}",
            "body": """Bonjour,

LA 27 Productions est un studio de musique haut de gamme basé à Barcelone. Nous composons de la musique 100% originale pour les marques, les agences et le cinéma — créée de toutes pièces avec des instruments réels, des enregistrements professionnels et sans banques de sons. Chaque pièce est exclusive à votre projet. Pas de royalties, pas de licences partagées, pas de musique que vos concurrents utilisent déjà.

Écoutez notre travail sur la27productions.com

Seriez-vous ouvert à un rapide appel si cela correspond à des projets en cours?

Tim Helmes"""
        }
    ],
    "pt": [
        {
            "subject": "Som para a próxima campanha da {company}",
            "body": """Olá,

LA 27 Productions é um estúdio de música premium com sede em Barcelona. Compomos música 100% original para marcas, agências e cinema — criada do zero com instrumentos reais, gravação profissional e sem bibliotecas de stock. Cada peça é exclusiva para o seu projeto. Sem royalties, sem licenças partilhadas, sem música que os seus concorrentes já estejam a usar.

Ouça o nosso trabalho em la27productions.com

Estaria aberto a uma chamada rápida se isso se encaixar em algum projeto?

Tim Helmes"""
        }
    ]
}

D2C_EMAIL_TEMPLATES = {
    "en": [
        {
            "subject": "Video & Audio for {company}",
            "body": """Hi,

We produce VSLs and video ads for health & supplement brands — with exclusive music and professional editing. All in-house from Barcelona.

Would you be open to a quick call if this fits any upcoming projects?

Tim Helmes
LA 27 Productions
la27productions.com"""
        }
    ],
    "es": [
        {
            "subject": "Vídeo & Audio para {company}",
            "body": """Hola,

Producimos VSLs y anuncios de vídeo para marcas de salud y suplementos, con música exclusiva y edición profesional. Todo in-house desde Barcelona.

¿Estarías abierto a una llamada rápida si esto encaja con algún proyecto?

Tim Helmes
LA 27 Productions
la27productions.com"""
        }
    ],
    "de": [
        {
            "subject": "Video & Audio für {company}",
            "body": """Hallo,

Wir produzieren VSLs and Video-Ads für Health & Supplement-Brands — mit exklusiver Musik und professionellem Schnitt. Alles in-house aus Barcelona.

Wären Sie offen für einen kurzen Call, falls das zu aktuellen Projekten passt?

Tim Helmes
LA 27 Productions
la27productions.com"""
        }
    ]
}

# Flat-file backup tracker — survives CRM locks
SENT_TRACKER = LOG_PATH / "sent_emails.json"

def _load_sent_tracker() -> set:
    """Load emails from flat-file tracker (never fails)."""
    try:
        if SENT_TRACKER.exists():
            with open(SENT_TRACKER, encoding="utf-8") as f:
                data = json.load(f)
            return set(e.lower().strip() for e in data)
    except Exception:
        pass
    return set()

def _save_sent_tracker(emails: set):
    """Persist sent emails to flat file."""
    try:
        LOG_PATH.mkdir(exist_ok=True)
        existing = _load_sent_tracker()
        merged = sorted(existing | emails)
        with open(SENT_TRACKER, "w", encoding="utf-8") as f:
            json.dump(merged, f, ensure_ascii=False)
    except Exception as e:
        print(f"  [Tracker] Could not save: {e}")

def get_contacted_emails() -> set:
    contacted = set()
    # Primary: read CRM xlsx
    try:
        wb = openpyxl.load_workbook(CRM_PATH)
        ws = wb["CRM"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3]:
                contacted.add(str(row[3]).lower().strip())
        print(f"  [CRM] Loaded {len(contacted)} contacted emails from CRM.")
    except Exception as e:
        print(f"  [CRM] Could not read CRM ({e}) — falling back to tracker file.")
    # Always merge with flat-file tracker (backup for when CRM is locked)
    tracker = _load_sent_tracker()
    if tracker:
        before = len(contacted)
        contacted |= tracker
        added = len(contacted) - before
        if added:
            print(f"  [Tracker] Added {added} emails from backup tracker.")
    return contacted


def add_to_crm(lead: dict, status: str = "ENVIADO", notes: str = ""):
    try:
        wb = openpyxl.load_workbook(CRM_PATH)
        ws = wb["CRM"]
        next_row = ws.max_row + 1
        row_num = next_row - 1
        ws.cell(row=next_row, column=1, value=row_num)
        ws.cell(row=next_row, column=2, value=lead.get("company", ""))
        ws.cell(row=next_row, column=3, value=lead.get("contact_name", ""))
        ws.cell(row=next_row, column=4, value=lead.get("email", ""))
        ws.cell(row=next_row, column=5, value=lead.get("location", ""))
        ws.cell(row=next_row, column=6, value=lead.get("industry", ""))
        ws.cell(row=next_row, column=7, value=datetime.date.today().isoformat())
        ws.cell(row=next_row, column=8, value=status)
        ws.cell(row=next_row, column=9, value=notes)
        wb.save(CRM_PATH)
        return True
    except Exception as e:
        print(f"  [CRM] Could not update CRM: {e}")
        return False


def detect_language(location: str, email: str = "", company: str = "", contact_name: str = "") -> str:
    """
    Detect language for outreach emails. v3 — FIXED.
    
    Priority: TLD → Location → Contact name → Company name → Default ENGLISH.
    
    KEY FIX: The old detector was too aggressive with Spanish matching,
    causing emails in Spanish to be sent to .co.uk, .com.au, .fi domains.
    Now we use STRICT matching and default to English for any ambiguity.
    """
    
    # === EXPLICIT TLD WHITELISTS (strongest signal) ===
    SPANISH_TLDS = {".es", ".mx", ".ar", ".cl", ".pe", ".ve",
                    ".ec", ".bo", ".uy", ".py", ".cr", ".pa", ".gt",
                    ".hn", ".ni", ".sv", ".cu", ".do", ".pr"}
    GERMAN_TLDS = {".de", ".at"}
    FRENCH_TLDS = {".fr"}
    PORTUGUESE_TLDS = {".pt", ".br"}
    # TLDs that are ENGLISH even if they look ambiguous
    ENGLISH_TLDS = {".com", ".co.uk", ".com.au", ".co.nz", ".us", ".uk", ".ca",
                    ".ie", ".org", ".net", ".io", ".ai", ".app", ".co",
                    ".fi", ".se", ".no", ".dk", ".nl", ".be", ".pl",
                    ".cz", ".sk", ".hu", ".ro", ".bg", ".hr", ".rs",
                    ".in", ".jp", ".kr", ".cn", ".sg", ".hk", ".tw",
                    ".za", ".ng", ".ke", ".eg", ".ae", ".sa", ".il",
                    ".ru", ".ua", ".tr", ".id", ".ph", ".th", ".vn", ".my"}
    
    # === LOCATION KEYWORDS ===
    SPANISH_GEO = {
        "spain", "españa", "espana", "madrid", "barcelona", "valencia",
        "sevilla", "seville", "bilbao", "malaga", "alicante", "murcia",
        "zaragoza", "palma", "las palmas", "vigo", "gijon", "cordoba",
        "mexico", "méxico", "cdmx", "guadalajara", "monterrey",
        "argentina", "buenos aires", "rosario",
        "colombia", "bogota", "bogotá", "medellin", "medellín",
        "chile", "santiago de chile", "peru", "lima", "venezuela", "caracas",
        "ecuador", "quito", "guayaquil", "bolivia", "la paz",
        "uruguay", "montevideo", "paraguay", "asuncion",
        "costa rica", "panama", "panamá", "guatemala",
        "honduras", "nicaragua", "el salvador", "cuba", "habana",
        "dominican republic", "santo domingo", "puerto rico",
    }
    GERMAN_GEO = {
        "germany", "deutschland", "austria", "österreich", "schweiz",
        "munich", "münchen", "berlin", "hamburg", "frankfurt",
        "cologne", "köln", "koln", "düsseldorf", "stuttgart", "vienna", "wien", "zürich", "zurich",
    }
    FRENCH_GEO = {
        "france", "paris", "lyon", "marseille", "bordeaux",
        "toulouse", "nantes", "strasbourg", "lille", "nice", "montpellier",
    }
    PORTUGUESE_GEO = {
        "portugal", "lisboa", "lisbon", "porto", "brazil", "brasil",
        "são paulo", "sao paulo", "rio de janeiro",
    }
    # Locations that should ALWAYS get English (even if contact has a Spanish name)
    ENGLISH_GEO = {
        "united states", "united kingdom", "australia", "canada",
        "new york", "los angeles", "london", "sydney", "melbourne",
        "toronto", "vancouver", "chicago", "san francisco",
        "seattle", "boston", "miami", "denver", "austin",
        "manchester", "birmingham", "dublin", "amsterdam",
        "singapore", "hong kong", "tokyo", "seoul", "mumbai",
        "delhi", "bangalore", "stockholm", "copenhagen", "oslo", "helsinki",
        "warsaw", "prague", "budapest", "bucharest",
    }
    
    # === SPANISH NAMES (only used when location is ambiguous) ===
    SPANISH_SURNAMES = {
        "garcia", "gonzalez", "hernandez", "lopez", "martinez", "moreno",
        "muñoz", "munoz", "navarro", "ramirez", "romero", "sanchez",
        "torres", "rodriguez", "ruiz", "jimenez", "gomez", "diaz",
        "fernandez", "alvarez", "reyes", "roig", "puig", "vidal",
        "fuster", "bosch", "ferrer",
    }
    
    email_l = email.lower().strip()
    company_l = company.lower().strip()
    location_l = location.lower().strip()
    contact_l = contact_name.lower().strip()
    
    # Extract domain parts
    domain = email_l.split("@")[-1] if "@" in email_l else ""
    
    # ============================================================
    # STEP 1: Check if location is EXPLICITLY English-speaking
    # If yes, ALWAYS send in English regardless of name/company
    # ============================================================
    for geo in ENGLISH_GEO:
        if geo in location_l:
            return "en"
    
    # ============================================================
    # STEP 2: TLD check (strongest signal after English geo)
    # ============================================================
    # Check ENGLISH TLDs first to prevent false Spanish matches on .co
    for tld in ENGLISH_TLDS:
        if domain.endswith(tld.lstrip(".")):
            # .co is tricky — could be Colombia or generic
            if tld == ".co" and not any(geo in location_l for geo in ["colombia", "bogota", "bogotá", "medellin", "medellín"]):
                pass  # Don't return yet, let location check decide
            elif tld == ".co":
                return "es"  # .co + Colombia location = Spanish
            # All other English TLDs: check location before defaulting
            # (a .com email from Madrid should get Spanish)
            break
    
    for tld in SPANISH_TLDS:
        if domain.endswith(tld.lstrip(".")):
            return "es"
    for tld in GERMAN_TLDS:
        if domain.endswith(tld.lstrip(".")):
            return "de"
    for tld in FRENCH_TLDS:
        if domain.endswith(tld.lstrip(".")):
            return "fr"
    for tld in PORTUGUESE_TLDS:
        if domain.endswith(tld.lstrip(".")):
            return "pt"
    
    # ============================================================
    # STEP 3: Location field (second strongest signal)
    # ============================================================
    for geo in SPANISH_GEO:
        if geo in location_l:
            return "es"
    for geo in GERMAN_GEO:
        if geo in location_l:
            return "de"
    for geo in FRENCH_GEO:
        if geo in location_l:
            return "fr"
    for geo in PORTUGUESE_GEO:
        if geo in location_l:
            return "pt"
    
    # ============================================================
    # STEP 4: Contact surname check (ONLY exact word match)
    # Only triggers if location didn't match anything
    # ============================================================
    if contact_l:
        contact_words = set(contact_l.split())
        if contact_words & SPANISH_SURNAMES:
            return "es"
    
    # ============================================================
    # STEP 5: DEFAULT = ENGLISH
    # When in doubt, English is safer — international business language
    # ============================================================
    return "en"


def is_on_ac_power() -> bool:
    try:
        import psutil
        battery = psutil.sensors_battery()
        if battery is None:
            return True
        return battery.power_plugged
    except Exception:
        return True


def get_site_snippet(website: str) -> str:
    if not website:
        return ""
    try:
        r = requests.get(website, timeout=6, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
        })
        if r.status_code != 200:
            return ""
        text = re.sub(r"<[^>]+>", " ", r.text)
        text = re.sub(r"\s+", " ", text).strip()
        return text[:400]
    except Exception:
        return ""


def ai_personalized_opener(lead: dict, lang: str) -> str:
    """Generate a personalized opening line using Ollama (free, local)."""
    # If Lead Finder already generated a perfect Icebreaker, use it immediately!
    pre_generated_icebreaker = lead.get("Icebreaker", "").strip()
    if pre_generated_icebreaker and len(pre_generated_icebreaker) > 10:
        return pre_generated_icebreaker

    ollama_url = os.getenv("OLLAMA_URL", "http://localhost:11434")
    ollama_model = os.getenv("OLLAMA_MODEL", "llama3.2:3b")
    try:
        company = lead.get("company", "")
        website = lead.get("website", "")
        site_text = get_site_snippet(website)
        context = site_text[:300] if site_text else ""

        # IMPORTANT: always specify LA 27 is a MUSIC studio, never video, never calls
        if lang == "es":
            prompt = (
                f"Eres Tim Helmes, productor musical de LA 27 Productions (estudio de MUSICA original para anuncios y marcas, Barcelona). "
                f"Escribe UNA sola frase de apertura personalizada para un email frio a '{company}'. "
                f"{'Contexto sobre la empresa: ' + context if context else ''} "
                "La frase debe hacer referencia a algo especifico de su sector o trabajo. "
                "Max 20 palabras. Sin saludo. Sin mencionar llamadas. Solo la frase."
            )
        elif lang == "de":
            prompt = (
                f"Du bist Tim Helmes, Musikproduzent von LA 27 Productions (Studio für ORIGINAL-MUSIK für Werbung und Marken, Barcelona). "
                f"Schreibe EINEN personalisierten Eröffnungssatz für eine Kalt-E-Mail an '{company}'. "
                f"{'Kontext: ' + context if context else ''} "
                "Klingt natürlich, max 20 Wörter, kein Gruß, keine Anrufe erwähnen. Nur der Satz."
            )
        else:
            prompt = (
                f"You are Tim Helmes, music producer at LA 27 Productions (studio for ORIGINAL MUSIC for ads and brands, Barcelona). "
                f"Write ONE personalized opening sentence for a cold email to '{company}'. "
                f"{'Context about them: ' + context if context else ''} "
                "Reference something specific about their work or industry. "
                "Max 20 words. No greeting. Do NOT mention calls or meetings. Just the sentence."
            )

        groq_api_key = os.getenv("GROQ_API_KEY")
        if not groq_api_key:
            return ""

        headers = {
            "Authorization": f"Bearer {groq_api_key}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": "llama3-8b-8192",
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.7
        }
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers=headers,
            json=payload,
            timeout=15,
        )
        if response.status_code == 200:
            opener = response.json().get("choices", [{}])[0].get("message", {}).get("content", "").strip()
            # Clean up — take only first sentence
            opener = opener.split(".")[0].strip()
            if opener and 5 < len(opener) < 200:
                return opener + "."
    except Exception as e:
        print(f"  [Groq] Opener error: {e}")
    return ""


def personalize_email(lead: dict, template: dict, use_ai: bool = True) -> tuple:
    company = lead.get("company", "your company")
    lang = lead.get("_lang", "en")

    subject = template["subject"].format(company=company)
    body = template["body"].format(company=company)

    if use_ai:
        opener = ai_personalized_opener(lead, lang)
        if opener:
            # Insert after the greeting line (first non-empty line)
            lines = body.split("\n")
            for i, line in enumerate(lines):
                if i > 0 and line.strip() == "":
                    lines.insert(i + 1, opener)
                    lines.insert(i + 2, "")
                    break
            body = "\n".join(lines)
            print(f"  [Ollama] Opener: {opener[:70]}...")

    return subject, body


def send_email_brevo(to_email: str, to_name: str, subject: str, body: str, lang: str = "en", account: str = "A") -> bool:
    """Send email via Brevo API. All emails go from tim@la27productions.com."""
    from email_template import build_html
    html_body = build_html(body, lang=lang)

    acc_upper = account.upper()
    if acc_upper == "A":
        key = BREVO_API_KEY_A
    elif acc_upper == "B":
        key = BREVO_API_KEY_B
    else:
        key = BREVO_API_KEY_C
        
    if not key:
        print(f"  [Send] ERROR: No API key for Brevo Account {account} configured")
        return False

    url = "https://api.api.brevo.com/v3/smtp/email" if "api.api." in key else "https://api.brevo.com/v3/smtp/email"
    headers = {
        "accept": "application/json",
        "api-key": key,
        "content-type": "application/json"
    }
    to_payload = {"email": to_email}
    if to_name and to_name.strip():
        to_payload["name"] = to_name.strip()

    payload = {
        "sender": {"name": SENDER_NAME, "email": SENDER_EMAIL},
        "to": [to_payload],
        "subject": subject,
        "textContent": body,
        "htmlContent": html_body
    }
    try:
        r = requests.post(url, headers=headers, json=payload, timeout=15)
        if r.status_code in (200, 201):
            return True
        else:
            print(f"  [Brevo Account {account}] Error {r.status_code}: {r.text[:200]}")
            return False
    except Exception as e:
        print(f"  [Brevo Account {account}] Exception: {e}")
        return False


def domain_is_valid(email: str) -> bool:
    """Check MX records and attempt SMTP handshake if possible, otherwise verify MX exists."""
    import socket
    import smtplib
    try:
        domain = email.split("@", 1)[1].strip()
    except Exception:
        return False
    
    # 1. Strict MX record verification
    try:
        import dns.resolver
        answers = dns.resolver.resolve(domain, 'MX')
        mx_hosts = sorted([(r.preference, str(r.exchange).strip()) for r in answers])
        if not mx_hosts:
            print(f"  [MX-Check] No MX records found for {domain}")
            return False
    except Exception as e:
        # If dns.resolver/DNS resolution fails, fallback to A-record DNS lookup to be safe
        print(f"  [MX-Check] MX DNS resolution failed for {domain} ({e}). Falling back to A-record.")
        try:
            socket.getaddrinfo(domain, None)
            return True
        except Exception:
            return False

    # 2. SMTP Handshake check (port 25)
    # If connection times out or gets refused (because GCP blocks port 25),
    # we catch the error and fallback to True (since MX records exist).
    # If the mail server explicitly rejects the recipient with a 5xx error, we return False!
    try:
        mx_host = mx_hosts[0][1]
        server = smtplib.SMTP(timeout=3)
        server.connect(mx_host, 25)
        server.helo("la27productions.com")
        server.mail("tim@la27productions.com")
        code, message = server.rcpt(email)
        server.quit()
        
        if code in (250, 251):
            return True
        elif code >= 500 and code < 600:
            print(f"  [SMTP-Check] Mailbox {email} REJECTED by MX server {mx_host}: Code {code} - {message}")
            return False
        else:
            return True
    except (socket.timeout, ConnectionRefusedError, OSError):
        # Socket timeout or connection refused (e.g. port 25 blocked outbound).
        # We fallback to True because the MX record exists and we cannot check the SMTP mailbox.
        return True
    except Exception:
        # Other errors (e.g. SMTP protocol error, SSL required, etc.)
        # Fallback to True since MX records exist.
        return True


def run_outreach(dry_run: bool = False) -> dict:
    print("  [Outreach] Starting outreach agent...")
    LOG_PATH.mkdir(exist_ok=True)

    if not LEADS_CSV.exists():
        print(f"  [Outreach] No leads CSV found at {LEADS_CSV}")
        return {"sent": 0, "skipped": 0, "error": "No leads file"}

    contacted = get_contacted_emails()
    print(f"  [Outreach] Already contacted: {len(contacted)} leads")

    # Use Ollama for AI personalization (free, local) — no API key needed
    ollama_url = os.getenv("OLLAMA_URL", "http://localhost:11434")
    try:
        _r = requests.get(f"{ollama_url}/api/tags", timeout=3)
        use_ai = _r.status_code == 200
    except Exception:
        use_ai = False
    if use_ai:
        print(f"  [Outreach] Ollama disponible — AI personalization ON")
    else:
        print(f"  [Outreach] Ollama no responde — usando templates fijos")

    sent = 0
    skipped = 0
    errors = 0
    sent_log = []
    sent_this_run = set()

    with open(LEADS_CSV, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if sent >= DAILY_LIMIT:
                print(f"  [Outreach] Daily limit ({DAILY_LIMIT}) reached.")
                break

            try:
                email = row.get("Email", "").replace("%20", "").strip().lower()
                if not email or "@" not in email:
                    skipped += 1
                    continue

                # Definitive email filters block
                prefix = email.split("@", 1)[0]
                domain = email.split("@", 1)[1] if "@" in email else ""

                # 1. Prefixes to block
                block_prefixes = [
                    "recruiting", "recrutement", "resume", "rfp", "rh", "hr", "register",
                    "rentals", "referral", "research", "representation", "subscriptions",
                    "sustainability", "tips", "trade", "upgrade", "visitors", "zendesk",
                    "windowscentral", "tecnologia", "tellmemore", "thinkglobal",
                    "datenschutz", "impressum", "news", "presse", "service",
                    "jobs", "careers", "interns", "recruitment", "hiring", "talent", "apply",
                    "press", "help", "logo", "zillow", "yourfriends", "compiled", "ihre"
                ]
                
                # Check if prefix matches or starts with blocked prefix (allowing exact or starts_with check)
                if any(prefix == p or prefix.startswith(p + ".") or prefix.startswith(p + "_") or prefix.startswith(p + "-") for p in block_prefixes):
                    print(f"  [Filter-Skip] Blocked prefix: {email}")
                    skipped += 1
                    continue

                # 2. Generic domains to block completely
                block_domains = [
                    "domain.com", "company.com", "example.com", "youragency.com", 
                    "yourbrand.com", "yourbusiness.com", "yourcompany.com", "acme.com", 
                    "dominio.com", "entreprise.com", "somewhere.com", "email.fr", 
                    "exemple.com", "doe.com", "domaine.com",
                    "dundermifflin.com", "test.com", "fake.com", "mailinator.com", "guerrillamail.com", "tempmail.com",
                    "email.de", "hospital.org", "logotyp.us", "gannett.com", "marcommnews.com", "militarytimes.com", "syracuse.com", "sior.com", "haymarket.co.in", "monsite.com"
                ]
                if domain in block_domains:
                    print(f"  [Filter-Skip] Blocked generic domain: {email}")
                    skipped += 1
                    continue

                # 3. Forbidden words/domains inside the email address
                block_words = [
                    "@gap.com", "@gov.", "@gc.ca", ".edu", "@brevo.com", 
                    "@zohocorp.com", "@pipedrive.com", "@adweek.com", "@fashionunited.com",
                    "jobs", "careers", "hiring", "recruitment", "licensing", "press", "redaccion", "autonews", "support", "assistenza", "hotro", "payroll",
                    "loudly", "ableton", "tunein", "lyst"
                ]
                if any(word in email for word in block_words):
                    print(f"  [Filter-Skip] Blocked forbidden word/domain: {email}")
                    skipped += 1
                    continue

                # 4. Large German/Austrian/DACH corporations and media
                large_corp_domains = [
                    "rtl.de", "sparkasse.at", "google.com", "axelspringer.de",
                    "rundfunkbeitrag.de", "otto.de", "amazon.de", "zalando.de",
                    "welt.de", "wiwo.de", "apotheken", "klinik", "krankenhaus",
                    "university", "uni-", "hochschule",
                    "universalproductionmusic.com", "theorchard.com", "hearst.com", "creativebloq.com", "futurenet.com", "wizzair.com", "aciertaretail.com", "umusic.com",
                    "loudly.com", "ableton.com", "tunein.com"
                ]
                if any(d in domain or d in email for d in large_corp_domains):
                    print(f"  [Filter-Skip] Blocked large corp/media domain: {email}")
                    skipped += 1
                    continue

                if email in contacted or email in sent_this_run:
                    skipped += 1
                    continue

                lead = {
                    "contact_name": row.get("Contact name", ""),
                    "email": email,
                    "company": row.get("Company name", ""),
                    "location": row.get("Location", ""),
                    "industry": row.get("Company industry", ""),
                    "title": row.get("Job title", ""),
                    "website": row.get("Company website", ""),
                }

                lang = detect_language(lead["location"], lead["email"], lead["company"], lead.get("contact_name", ""))
                lead["_lang"] = lang
                import random as _rnd
                industry_lower = lead.get("industry", "").lower()
                is_d2c = "supplement" in industry_lower or "nutrition" in industry_lower
                
                if is_d2c:
                    templates_list = D2C_EMAIL_TEMPLATES.get(lang, D2C_EMAIL_TEMPLATES["en"])
                else:
                    templates_list = EMAIL_TEMPLATES.get(lang, EMAIL_TEMPLATES["en"])
                    
                template = _rnd.choice(templates_list)
                subject, body = personalize_email(lead, template, use_ai=use_ai)

                if dry_run:
                    print(f"  [DRY RUN] Would send to: {email} ({lead['company']}) [{lang.upper()}]")
                    sent += 1
                    continue

                # Determine which account has quota and alternate them
                import quota_manager
                
                can_a = quota_manager.can_send("A")
                can_b = quota_manager.can_send("B")
                can_c = quota_manager.can_send("C")

                available_accounts = []
                if can_a: available_accounts.append("A")
                if can_b: available_accounts.append("B")
                if can_c: available_accounts.append("C")

                if not available_accounts:
                    print("  [Outreach] All Brevo accounts (A, B & C) have reached their daily limits (290/290 each).")
                    break

                # Validate domain before sending (avoids bounce + protects reputation)
                if not domain_is_valid(email):
                    print(f"  [Skip] Invalid domain: {email}")
                    skipped += 1
                    continue

                success = False
                while available_accounts and not success:
                    if len(available_accounts) == 1:
                        target_account = available_accounts[0]
                    else:
                        # Select the account with the minimum sent count among those available
                        try:
                            with open(quota_manager.QUOTA_FILE, "r") as qf:
                                qdata = json.load(qf)
                            usage = {
                                "A": qdata.get("sent_a", 0),
                                "B": qdata.get("sent_b", 0),
                                "C": qdata.get("sent_c", 0)
                            }
                            available_usage = {k: v for k, v in usage.items() if k in available_accounts}
                            target_account = min(available_usage, key=available_usage.get)
                        except Exception:
                            target_account = available_accounts[0]

                    print(f"  [Outreach] Sending to {email} ({lead['company']}) [{lang.upper()}] via Account {target_account}...")
                    success = send_email_brevo(email, lead["contact_name"], subject, body, lang=lang, account=target_account)

                    if success:
                        sent += 1
                        sent_this_run.add(email)
                        quota_manager.increment_quota(target_account)
                        add_to_crm(lead, status="ENVIADO", notes=f"Cold email LANG:{lang.upper()} Account:{target_account} — {datetime.date.today()}")
                        _save_sent_tracker(sent_this_run)
                        sent_log.append({"email": email, "company": lead["company"], "lang": lang, "account": target_account})
                    else:
                        errors += 1
                        print(f"  [Outreach] Account {target_account} failed to send. Removing from available accounts for this run.")
                        available_accounts.remove(target_account)

                if not success:
                    print(f"  [Outreach] Failed to send email to {email} using any of the available accounts.")
                    continue

                # Wait between every sending attempt to prevent hammering and respect delay settings
                import random as _r
                time.sleep(_r.uniform(45, 180))
            except Exception as e:
                print(f"  [Outreach] Exception processing lead {row.get('Email', '')}: {e}")
                errors += 1
                continue

    log_file = LOG_PATH / f"outreach_{datetime.date.today()}.json"
    with open(log_file, "w", encoding="utf-8") as f:
        json.dump(
            {"date": datetime.date.today().isoformat(), "sent": sent,
             "skipped": skipped, "errors": errors, "leads": sent_log},
            f, indent=2, ensure_ascii=False, default=str
        )

    print(f"  [Outreach] Done. Sent: {sent}, Skipped: {skipped}, Errors: {errors}")
    return {"sent": sent, "skipped": skipped, "errors": errors}


if __name__ == "__main__":
    import json
    result = run_outreach(dry_run=True)
    print(json.dumps(result, indent=2))
