import openpyxl
import csv
from datetime import date

crm_path = "/home/timhe/la27/LA27_CRM.xlsx"
leads_csv_path = "/home/timhe/la27/LA27_leads_with_email.csv"

# 1. Build email to website map from leads CSV
email_to_url = {}
try:
    with open(leads_csv_path, "r", encoding="utf-8", errors="ignore") as f:
        reader = csv.DictReader(f)
        for row in reader:
            email = row.get("Email", "").strip().lower()
            url = row.get("Company website", "").strip()
            if email:
                email_to_url[email] = url
except Exception as e:
    print(f"Error reading leads CSV: {e}")

# 2. Read CRM for today's leads
today_str = date.today().isoformat() # '2026-06-16'
agencies = []
supplements = []

try:
    wb = openpyxl.load_workbook(crm_path)
    ws = wb["CRM"]
    for row in range(2, ws.max_row + 1):
        vals = [cell.value for cell in ws[row]]
        # Column 7 is FECHA
        row_date = vals[6]
        if str(row_date).startswith(today_str):
            company = vals[1]
            email = str(vals[3]).strip().lower()
            tipo = vals[5]
            url = email_to_url.get(email, "No URL found")
            
            lead_info = {
                "company": company,
                "email": email,
                "url": url,
                "type": tipo
            }
            
            if tipo == "D2C Nutrition / Health Supplements":
                supplements.append(lead_info)
            else:
                agencies.append(lead_info)
except Exception as e:
    print(f"Error reading CRM: {e}")

# Generate Markdown file
md_content = f"""# Leads Encontrados Hoy (16 de Junio de 2026)

Este reporte contiene los leads recopilados hoy por los agentes de LA 27 Productions, organizados por su flujo de negocio.

## Resumen
* **Flujo Agencias (Premium Advertising):** {len(agencies)} leads
* **Flujo Suplementos Alemanes (D2C):** {len(supplements)} leads

---

## Flujo Agencias (Premium Advertising)
| # | Empresa | Email | URL |
|---|---------|-------|-----|
"""

for idx, lead in enumerate(agencies):
    md_content += f"| {idx+1} | {lead['company']} | {lead['email']} | [{lead['url']}]({lead['url']}) |\n"

md_content += """
---

## Flujo Suplementos Alemanes (D2C)
| # | Empresa | Email | URL |
|---|---------|-------|-----|
"""

for idx, lead in enumerate(supplements):
    md_content += f"| {idx+1} | {lead['company']} | {lead['email']} | [{lead['url']}]({lead['url']}) |\n"

with open("/home/timhe/la27/agents/logs/today_leads_report.md", "w", encoding="utf-8") as f:
    f.write(md_content)

print("Markdown report generated at agents/logs/today_leads_report.md")
